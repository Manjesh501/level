from pdfminer.high_level import extract_text as extract_pdf_text
from docx import Document
import openpyxl

def extract_text_from_pdf(file_path):
    return extract_pdf_text(file_path)

def extract_text_from_docx(file_path):
    doc = Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def extract_text_from_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    all_text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    all_text += str(cell.value) + " "
    return all_text
